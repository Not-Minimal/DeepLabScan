#!/usr/bin/env python3
"""
Excel Logger para DeepLabScan

Este módulo maneja el guardado de resultados de entrenamiento, evaluación y predicción
en un archivo Excel compartido con múltiples hojas:
- Hoja "Resumen": Todos los experimentos juntos para comparación
- Hoja "Training": Resultados de entrenamiento
- Hoja "Evaluation": Resultados de evaluación
- Hoja "Prediction": Resultados de predicción

Uso:
    from scripts.excel_logger import ExcelLogger

    logger = ExcelLogger()
    logger.log_training(results_dict)
    logger.log_evaluation(results_dict)
    logger.log_prediction(results_dict)
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: Se requieren pandas y openpyxl")
    print("Instala con: pip install pandas openpyxl")
    exit(1)


class ExcelLogger:
    """
    Clase para manejar el logging de resultados en Excel
    """

    def __init__(self, excel_path: str = "results/experiment_results.xlsx"):
        """
        Inicializa el logger de Excel

        Args:
            excel_path: Ruta al archivo Excel donde se guardarán los resultados
        """
        self.excel_path = Path(excel_path)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)

        # Nombres de las hojas
        self.SUMMARY_SHEET = "Resumen"
        self.TRAIN_SHEET = "Training"
        self.EVAL_SHEET = "Evaluation"
        self.PREDICT_SHEET = "Prediction"

        # Inicializar archivo si no existe
        self._initialize_excel()

    def _initialize_excel(self):
        """
        Crea el archivo Excel con las hojas necesarias si no existe
        """
        if not self.excel_path.exists():
            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                # Crear hojas vacías con encabezados

                # Hoja Resumen
                summary_columns = [
                    "Fecha",
                    "Hora",
                    "Tipo",
                    "Experimento",
                    "Modelo",
                    "Épocas",
                    "mAP@0.5",
                    "mAP@0.5:0.95",
                    "Precision",
                    "Recall",
                    "F1-Score",
                    "Detecciones",
                    "Notas",
                ]
                pd.DataFrame(columns=summary_columns).to_excel(
                    writer, sheet_name=self.SUMMARY_SHEET, index=False
                )

                # Hoja Training
                train_columns = [
                    "Fecha",
                    "Hora",
                    "Experimento",
                    "Modelo",
                    "Dataset",
                    "Épocas",
                    "Batch",
                    "Imgsz",
                    "Dispositivo",
                    "Duración (min)",
                    "Best mAP@0.5",
                    "Best mAP@0.5:0.95",
                    "Best Precision",
                    "Best Recall",
                    "Final Loss",
                    "Weights Path",
                    "Notas",
                ]
                pd.DataFrame(columns=train_columns).to_excel(
                    writer, sheet_name=self.TRAIN_SHEET, index=False
                )

                # Hoja Evaluation
                eval_columns = [
                    "Fecha",
                    "Hora",
                    "Experimento",
                    "Weights",
                    "Dataset",
                    "Split",
                    "Dispositivo",
                    "Precision",
                    "Recall",
                    "mAP@0.5",
                    "mAP@0.5:0.95",
                    "F1-Score",
                    "Clases Detectadas",
                    "Visualizaciones",
                    "Notas",
                ]
                pd.DataFrame(columns=eval_columns).to_excel(
                    writer, sheet_name=self.EVAL_SHEET, index=False
                )

                # Hoja Prediction
                predict_columns = [
                    "Fecha",
                    "Hora",
                    "Experimento",
                    "Weights",
                    "Source",
                    "Confidence",
                    "IoU",
                    "Dispositivo",
                    "Total Imágenes",
                    "Total Detecciones",
                    "Detecciones por Clase",
                    "Output Dir",
                    "Notas",
                ]
                pd.DataFrame(columns=predict_columns).to_excel(
                    writer, sheet_name=self.PREDICT_SHEET, index=False
                )

            # Aplicar formato
            self._format_excel()

            print(f"✓ Archivo Excel inicializado: {self.excel_path}")

    def _format_excel(self):
        """
        Aplica formato visual al archivo Excel
        """
        try:
            wb = load_workbook(self.excel_path)

            # Estilo para encabezados
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Formatear encabezados
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.excel_path)
        except Exception as e:
            print(f"⚠️  No se pudo aplicar formato: {e}")

    def _append_row(self, sheet_name: str, data: Dict[str, Any]):
        """
        Agrega una fila a una hoja específica

        Args:
            sheet_name: Nombre de la hoja
            data: Diccionario con los datos a agregar
        """
        try:
            # Leer hoja existente
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            # Agregar nueva fila
            new_row = pd.DataFrame([data])
            df = pd.concat([df, new_row], ignore_index=True)

            # Guardar
            with pd.ExcelWriter(
                self.excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Reformatear
            self._format_excel()

            return True
        except Exception as e:
            print(f"❌ Error al agregar fila a {sheet_name}: {e}")
            return False

    def log_training(
        self,
        experiment_name: str,
        model: str,
        dataset: str,
        epochs: int,
        batch: int,
        imgsz: int,
        device: str,
        duration_minutes: float,
        best_map50: float,
        best_map50_95: float,
        best_precision: float,
        best_recall: float,
        final_loss: float,
        weights_path: str,
        notes: str = "",
    ) -> bool:
        """
        Registra resultados de entrenamiento
        """
        now = datetime.now()

        # Calcular F1-Score
        f1_score = 0.0
        if best_precision > 0 or best_recall > 0:
            f1_score = (
                2 * (best_precision * best_recall) / (best_precision + best_recall)
                if (best_precision + best_recall) > 0
                else 0
            )

        # Datos para hoja Training
        train_data = {
            "Fecha": now.strftime("%Y-%m-%d"),
            "Hora": now.strftime("%H:%M:%S"),
            "Experimento": experiment_name,
            "Modelo": model,
            "Dataset": dataset,
            "Épocas": epochs,
            "Batch": batch,
            "Imgsz": imgsz,
            "Dispositivo": device,
            "Duración (min)": round(duration_minutes, 2),
            "Best mAP@0.5": round(best_map50, 4),
            "Best mAP@0.5:0.95": round(best_map50_95, 4),
            "Best Precision": round(best_precision, 4),
            "Best Recall": round(best_recall, 4),
            "Final Loss": round(final_loss, 4),
            "Weights Path": weights_path,
            "Notas": notes,
        }

        # Datos para hoja Resumen
        summary_data = {
            "Fecha": now.strftime("%Y-%m-%d"),
            "Hora": now.strftime("%H:%M:%S"),
            "Tipo": "Training",
            "Experimento": experiment_name,
            "Modelo": model,
            "Épocas": epochs,
            "mAP@0.5": round(best_map50, 4),
            "mAP@0.5:0.95": round(best_map50_95, 4),
            "Precision": round(best_precision, 4),
            "Recall": round(best_recall, 4),
            "F1-Score": round(f1_score, 4),
            "Detecciones": "-",
            "Notas": notes,
        }

        # Agregar a ambas hojas
        success_train = self._append_row(self.TRAIN_SHEET, train_data)
        success_summary = self._append_row(self.SUMMARY_SHEET, summary_data)

        if success_train and success_summary:
            print(f"✓ Resultados de entrenamiento guardados en {self.excel_path}")
            return True
        return False

    def log_evaluation(
        self,
        experiment_name: str,
        weights_path: str,
        dataset: str,
        split: str,
        device: str,
        precision: float,
        recall: float,
        map50: float,
        map50_95: float,
        classes_detected: int,
        visualizations_count: int = 0,
        notes: str = "",
    ) -> bool:
        """
        Registra resultados de evaluación
        """
        now = datetime.now()

        # Calcular F1-Score
        f1_score = 0.0
        if precision > 0 or recall > 0:
            f1_score = (
                2 * (precision * recall) / (precision + recall)
                if (precision + recall) > 0
                else 0
            )

        # Datos para hoja Evaluation
        eval_data = {
            "Fecha": now.strftime("%Y-%m-%d"),
            "Hora": now.strftime("%H:%M:%S"),
            "Experimento": experiment_name,
            "Weights": weights_path,
            "Dataset": dataset,
            "Split": split,
            "Dispositivo": device,
            "Precision": round(precision, 4),
            "Recall": round(recall, 4),
            "mAP@0.5": round(map50, 4),
            "mAP@0.5:0.95": round(map50_95, 4),
            "F1-Score": round(f1_score, 4),
            "Clases Detectadas": classes_detected,
            "Visualizaciones": visualizations_count,
            "Notas": notes,
        }

        # Datos para hoja Resumen
        summary_data = {
            "Fecha": now.strftime("%Y-%m-%d"),
            "Hora": now.strftime("%H:%M:%S"),
            "Tipo": "Evaluation",
            "Experimento": experiment_name,
            "Modelo": Path(weights_path).parent.parent.name,
            "Épocas": "-",
            "mAP@0.5": round(map50, 4),
            "mAP@0.5:0.95": round(map50_95, 4),
            "Precision": round(precision, 4),
            "Recall": round(recall, 4),
            "F1-Score": round(f1_score, 4),
            "Detecciones": "-",
            "Notas": notes,
        }

        # Agregar a ambas hojas
        success_eval = self._append_row(self.EVAL_SHEET, eval_data)
        success_summary = self._append_row(self.SUMMARY_SHEET, summary_data)

        if success_eval and success_summary:
            print(f"✓ Resultados de evaluación guardados en {self.excel_path}")
            return True
        return False

    def log_prediction(
        self,
        experiment_name: str,
        weights_path: str,
        source: str,
        confidence: float,
        iou: float,
        device: str,
        total_images: int,
        total_detections: int,
        class_counts: Dict[str, int],
        output_dir: str,
        notes: str = "",
    ) -> bool:
        """
        Registra resultados de predicción
        """
        now = datetime.now()

        # Formatear detecciones por clase
        class_counts_str = ", ".join(
            [f"{cls}: {count}" for cls, count in class_counts.items()]
        )
        if not class_counts_str:
            class_counts_str = "Sin detecciones"

        # Datos para hoja Prediction
        predict_data = {
            "Fecha": now.strftime("%Y-%m-%d"),
            "Hora": now.strftime("%H:%M:%S"),
            "Experimento": experiment_name,
            "Weights": weights_path,
            "Source": source,
            "Confidence": confidence,
            "IoU": iou,
            "Dispositivo": device,
            "Total Imágenes": total_images,
            "Total Detecciones": total_detections,
            "Detecciones por Clase": class_counts_str,
            "Output Dir": output_dir,
            "Notas": notes,
        }

        # Datos para hoja Resumen
        summary_data = {
            "Fecha": now.strftime("%Y-%m-%d"),
            "Hora": now.strftime("%H:%M:%S"),
            "Tipo": "Prediction",
            "Experimento": experiment_name,
            "Modelo": Path(weights_path).parent.parent.name,
            "Épocas": "-",
            "mAP@0.5": "-",
            "mAP@0.5:0.95": "-",
            "Precision": "-",
            "Recall": "-",
            "F1-Score": "-",
            "Detecciones": total_detections,
            "Notas": f"{class_counts_str} | {notes}" if notes else class_counts_str,
        }

        # Agregar a ambas hojas
        success_predict = self._append_row(self.PREDICT_SHEET, predict_data)
        success_summary = self._append_row(self.SUMMARY_SHEET, summary_data)

        if success_predict and success_summary:
            print(f"✓ Resultados de predicción guardados en {self.excel_path}")
            return True
        return False

    def get_summary_dataframe(self) -> pd.DataFrame:
        """
        Retorna un DataFrame con el resumen de todos los experimentos
        """
        try:
            return pd.read_excel(self.excel_path, sheet_name=self.SUMMARY_SHEET)
        except Exception as e:
            print(f"❌ Error al leer resumen: {e}")
            return pd.DataFrame()

    def get_best_model(self, metric: str = "mAP@0.5") -> Optional[Dict[str, Any]]:
        """
        Retorna el mejor modelo según una métrica específica

        Args:
            metric: Métrica a usar ('mAP@0.5', 'mAP@0.5:0.95', 'F1-Score', etc.)
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name=self.SUMMARY_SHEET)

            # Filtrar solo entrenamientos y evaluaciones
            df_filtered = df[df["Tipo"].isin(["Training", "Evaluation"])]

            # Convertir métrica a numérico
            df_filtered[metric] = pd.to_numeric(df_filtered[metric], errors="coerce")

            # Encontrar el mejor
            best_row = df_filtered.loc[df_filtered[metric].idxmax()]

            return best_row.to_dict()
        except Exception as e:
            print(f"❌ Error al buscar mejor modelo: {e}")
            return None

    def print_summary(self, last_n: int = 10):
        """
        Imprime un resumen de los últimos experimentos

        Args:
            last_n: Número de últimos experimentos a mostrar
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name=self.SUMMARY_SHEET)

            if len(df) == 0:
                print("No hay experimentos registrados todavía.")
                return

            print("\n" + "=" * 80)
            print(f"RESUMEN DE ÚLTIMOS {min(last_n, len(df))} EXPERIMENTOS")
            print("=" * 80)

            # Mostrar últimos n experimentos
            recent_df = df.tail(last_n)

            # Configurar pandas para mostrar todas las columnas
            pd.set_option("display.max_columns", None)
            pd.set_option("display.width", None)
            pd.set_option("display.max_colwidth", 30)

            print(recent_df.to_string(index=False))
            print("=" * 80)

        except Exception as e:
            print(f"❌ Error al imprimir resumen: {e}")


# Función de utilidad para usar desde los scripts
def get_logger(excel_path: str = "results/experiment_results.xlsx") -> ExcelLogger:
    """
    Retorna una instancia del logger de Excel
    """
    return ExcelLogger(excel_path)


if __name__ == "__main__":
    # Test del logger
    print("Probando ExcelLogger...")

    logger = ExcelLogger("results/test_results.xlsx")

    # Test training
    logger.log_training(
        experiment_name="test_exp_1",
        model="yolo11n.pt",
        dataset="data/raw",
        epochs=15,
        batch=16,
        imgsz=640,
        device="cuda",
        duration_minutes=25.5,
        best_map50=0.85,
        best_map50_95=0.65,
        best_precision=0.82,
        best_recall=0.78,
        final_loss=0.15,
        weights_path="runs/detect/train/weights/best.pt",
        notes="Experimento de prueba",
    )

    print("\n✓ Test completado. Verifica results/test_results.xlsx")
